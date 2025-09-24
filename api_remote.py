#!/home/sami/santander/venv/bin/python

import logging
import re
import os
import sys
import time
import random
from datetime import datetime

# Registrar inicio del script
print("=== scrapeo.py iniciado ===")

# Modo opcional: solo importar Excel y salir, controlado por la variable de entorno ONLY_IMPORT
def _import_only():
    try:
        import os as _os
        from pathlib import Path as _Path
        env_path = _os.environ.get('DOWNLOAD_PATH')
        posibles = []
        if env_path:
            p = _Path(env_path)
            posibles.extend([p, p.with_suffix('.xlsx')])
        posibles.extend([
            _Path('/tmp/descarga.xlsx'),
            _Path('/tmp/descarga'),
            _Path.cwd() / 'descarga.xlsx',
            _Path.cwd() / 'descarga'
        ])
        file_path = next((p for p in posibles if p.exists()), None)
        if not file_path:
            print("No se encontró el archivo de descarga (use DOWNLOAD_PATH o /tmp/descarga.xlsx)")
            sys.exit(1)
        if file_path.is_file() and file_path.stat().st_size == 0:
            print(f"El archivo {file_path} está vacío")
            sys.exit(1)
        try:
            import pandas as _pd
            df = _pd.read_excel(str(file_path))
            print(f"Importación correcta: filas={len(df)}, columnas={len(df.columns)}")
            try:
                cols = list(map(str, list(df.columns)[:10]))
                if cols:
                    print("Columnas:", ", ".join(cols))
            except Exception:
                pass
            sys.exit(0)
        except Exception:
            pass
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)
            ws = wb.active
            n_rows = ws.max_row or 0
            n_cols = ws.max_column or 0
            print(f"Importación correcta (openpyxl): filas~={n_rows}, columnas~={n_cols}")
            try:
                headers = []
                for c in range(1, min(n_cols, 10) + 1):
                    v = ws.cell(row=1, column=c).value
                    headers.append(str(v) if v is not None else '')
                if headers:
                    print("Columnas:", ", ".join(headers))
            except Exception:
                pass
            sys.exit(0)
        except Exception as _eox:
            print(f"No se pudo importar el Excel: {_eox}")
            sys.exit(1)
    except Exception as _eimp:
        print(f"Error en importación: {_eimp}")
        sys.exit(1)

if os.environ.get('ONLY_IMPORT', '').lower() in ('1','true','yes','on'):
    _import_only()

# Verificar dependencias
try:
    from playwright.sync_api import sync_playwright
    import os
    print("✓ Playwright importado correctamente")
except ImportError:
    print("✗ ERROR: Playwright no está instalado. Intentando instalar...")
    try:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "playwright"])
        subprocess.check_call([sys.executable, "-m", "playwright", "install", "chromium"])
        from playwright.sync_api import sync_playwright
        print("✓ Playwright instalado correctamente")
    except Exception as e:
        print(f"✗ ERROR: No se pudo instalar Playwright: {e}")
        sys.exit(1)

# Deshabilitado: no se generan logs ni evidencias
log_dir = None

# Deshabilitar logging a fichero y a stdout (sin logs)
logging.disable(logging.CRITICAL)

# from db_utils import get_db_connection  # Eliminado: ya no se realiza conexión a la BD

MAX_LOGIN_TIME = 60  # segundos
nif = "44007535W"
clave = "19731898"

fecha_inicio = datetime.now().replace(day=1).strftime("%d/%m/%Y")
#fecha_inicio = "01/01/2025"
fecha_hoy = datetime.now().strftime("%d/%m/%Y")

# Guardar mes y año del filtro para comparación posterior
mes_filtro = int(fecha_inicio[3:5])
anio_filtro = int(fecha_inicio[6:10])
mes_anio = f"{str(mes_filtro).zfill(2)}/{anio_filtro}"

with sync_playwright() as p:
    try:
        # Configuración para evitar detección anti-bot más avanzada
        browser_args = [
            '--disable-blink-features=AutomationControlled',
            '--no-sandbox',
            '--disable-setuid-sandbox',
            '--disable-dev-shm-usage',
            '--disable-accelerated-2d-canvas',
            '--no-first-run',
            '--no-zygote',
            '--disable-gpu',
            '--window-size=1920,1080',
            '--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            '--disable-extensions',
            '--disable-features=IsolateOrigins,site-per-process',
            '--enable-features=NetworkService',
            '--ignore-certificate-errors',
            '--disable-web-security',
            '--allow-running-insecure-content',
            '--disable-webgl'
        ]
        
        # Configuración para control de visibilidad del navegador
        show_browser = os.environ.get('SHOW_BROWSER', '0').lower() in ('1', 'true', 'yes', 'on')
        
        # Forzar headless si se especifica en el entorno
        if os.environ.get('HEADLESS', '').lower() in ('1', 'true', 'yes'):
            show_browser = False
        # Permitir forzar modo NO persistente por variable de entorno (workaround para 192.168.1.55)
        force_non_persist = os.environ.get('PW_NO_PERSIST', '').lower() in ('1', 'true', 'yes', 'on')
        if force_non_persist:
            logging.info("PW_NO_PERSIST=1 -> Usando navegador NO persistente")
            browser = p.chromium.launch(headless=not show_browser, args=browser_args, slow_mo=100 if show_browser else 0)
            context = browser.new_context(
                viewport={'width': 1400, 'height': 900},
                locale='es-ES'
            )
        else:
            # Directorio de perfil persistente para reutilizar cookies/sesión y reducir bloqueos de WAF
            user_data_dir = "/var/www/html/.pw_profile_santander"
            os.makedirs(user_data_dir, exist_ok=True)
            # Lanzar contexto persistente directamente (evita crear new_context luego)
            try:
                context = p.chromium.launch_persistent_context(
                    user_data_dir=user_data_dir,
                    headless=not show_browser,
                    args=browser_args,
                    slow_mo=100 if show_browser else 0,
                    accept_downloads=True,
                    viewport={'width': 1920, 'height': 1080},
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                    locale="es-ES",
                    timezone_id="Europe/Madrid",
                    permissions=['geolocation', 'notifications'],
                    extra_http_headers={
                        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
                        "Accept-Language": "es-ES,es;q=0.9,en-US;q=0.8,en;q=0.7",
                        "Cache-Control": "max-age=0",
                        "Connection": "keep-alive",
                        "Sec-Fetch-Dest": "document",
                        "Sec-Fetch-Mode": "navigate",
                        "Sec-Fetch-Site": "same-origin",
                        "Sec-Fetch-User": "?1"
                    }
                )
                logging.info("Contexto persistente iniciado (perfil reutilizable) con configuración anti-detección avanzada")
            except Exception as e_persist:
                logging.warning(f"No se pudo iniciar contexto persistente, usando navegador no persistente: {e_persist}")
                # Fallback a navegador no persistente manteniendo visibilidad
                browser = p.chromium.launch(headless=not show_browser, args=browser_args, slow_mo=100 if show_browser else 0)
                context = browser.new_context(
                    viewport={'width': 1400, 'height': 900},
                    locale='es-ES'
                )
    except Exception as e:
        logging.error(f"Error al iniciar el navegador: {e}")
        sys.exit(1)
        
    # Contexto ya creado arriba con perfil persistente
    # Aumentar timeout por defecto para todas las operaciones del contexto
    context.set_default_timeout(30000)
    # Iniciar tracing para depurar interacciones y cierres de página
    try:
        context.tracing.start(screenshots=True, snapshots=True, sources=True)
    except Exception as e:
        logging.warning(f"No se pudo iniciar tracing de Playwright: {e}")
    
    # Inyectar JavaScript para evadir detección de bots (overriding navigator properties)
    context.add_init_script('''
    Object.defineProperty(navigator, 'webdriver', { get: () => false });    
    Object.defineProperty(navigator, 'languages', { get: () => ['es-ES', 'es', 'en-US', 'en'] });    
    Object.defineProperty(navigator, 'plugins', { get: () => Array(3).fill().map((_, i) => i) });    
    window.chrome = { runtime: {} };
    ''')
    
    # Crear página con configuración avanzada
    page = context.new_page()
    # Bloqueo de navegación a nivel de contexto y página (red, eventos y navegación)
    try:
        # 1) Bloqueo de red por contexto (regex)
        context.route(re.compile(r"asuntos-clientes", re.I), lambda route: route.abort())
    except Exception:
        pass
    try:
        # 2) Cancelar clics en anchors con href prohibidos
        page.add_init_script(
            """
            (()=>{
              const bad = /asuntos-clientes|buzon|ayuda|gestiones/i;
              window.addEventListener('click', (ev)=>{
                try{
                  const a = ev.target.closest && ev.target.closest('a[href]');
                  if (a){
                    const href = a.getAttribute('href')||'';
                    if (bad.test(href)){
                      ev.preventDefault(); ev.stopPropagation(); return false;
                    }
                  }
                }catch(e){}
              }, true);
            })();
            """
        )
    except Exception:
        pass
    try:
        # 3) Observador: si se navega fuera del detalle o a asuntos-clientes, volver atrás
        def _nav_guard(frame):
            try:
                u = frame.url or ''
            except Exception:
                u = ''
            try:
                if re.search(r"asuntos-clientes", u, re.I):
                    logging.warning(f"Intercepción de navegación a {u}; volviendo atrás")
                    try:
                        page.go_back()
                    except Exception:
                        pass
                elif u and not re.search(r"/oneweb/nhb/cuentas/detalle", u, re.I):
                    # Permitimos dominios intermedios si no hay cambio sustancial
                    pass
            except Exception:
                pass
        # 2.c) Enumeración de candidatos en barras de acción (con evidencias) y dentro de 'paesfr-transactions'
        if not dl_clicked_pre:
            try:
                bar_sel = ".buttons, .actions, .toolbar, .san-actions, .san-toolbar, .table-actions, .actions-bar, [role='toolbar']"
                cand_sel = f"{bar_sel} >> :is(button,a,[role=button],[aria-label],[title])"
                scopes = [scope]
                try:
                    # Incluir web component si existe
                    wcomp = scope.locator("paesfr-transactions").first
                    if wcomp.count() > 0:
                        scopes.append(wcomp)
                except Exception:
                    pass
                for sc in scopes:
                    loc = sc.locator(cand_sel)
                    cnt = 0
                    try:
                        total = loc.count()
                    except Exception:
                        total = 0
                    for i in range(min(total, 12)):
                        try:
                            el = loc.nth(i)
                            # Extraer metadatos del candidato
                            meta = {}
                            try:
                                meta = el.evaluate('el => ({\n'+
                                                   '  text: (el.innerText||"").trim(),\n'+
                                                   '  aria: el.getAttribute && (el.getAttribute("aria-label")||""),\n'+
                                                   '  title: el.getAttribute && (el.getAttribute("title")||""),\n'+
                                                   '  cls: el.className||"",\n'+
                                                   '  hasSvg: !!el.querySelector("svg"),\n'+
                                                   '  href: (el.tagName && el.tagName.toLowerCase()==="a") ? (el.getAttribute("href")||"") : (el.closest("a") ? (el.closest("a").getAttribute("href")||"") : "")\n'+
                                                   '})')
                            except Exception:
                                meta = {}
                            txt = (meta.get('text') or '').strip()
                            aria = meta.get('aria') or ''
                            titl = meta.get('title') or ''
                            href = meta.get('href') or ''
                            # Heurísticas de descarga/exportación
                            looks_download = bool(re.search(r"descarg|export", (txt + ' ' + aria + ' ' + titl), re.I)) or (len(txt) <= 2 and (meta.get('hasSvg') or False))
                            if not looks_download:
                                continue
                            # Validaciones anti-header/menú
                            ok = True
                            try:
                                ok = el.evaluate(
                                    "el => {\n"+
                                    "  const inBadAnc = !!el.closest('header, nav, san-sidebar-menu, #personalAreaMenu, #mailboxMenu, #helpMenu, .menu, [role=\\\"menu\\\"]');\n"+
                                    "  if (inBadAnc) return false;\n"+
                                    "  const a = (el.tagName && el.tagName.toLowerCase()==='a') ? el : el.closest('a');\n"+
                                    "  const href = a && a.getAttribute && (a.getAttribute('href')||'');\n"+
                                    "  if (href && /asuntos-clientes|buzon|ayuda|gestiones/i.test(href)) return false;\n"+
                                    "  return true;\n"+
                                    "}"
                                )
                            except Exception:
                                ok = True
                            if not ok:
                                continue
                            # Evidencias deshabilitadas
                            # Sin capturas PNG ni intentos de clic; solo guardamos HTML del candidato
                            try:
                                el.scroll_into_view_if_needed()
                            except Exception:
                                pass
                            # No interactuar: continuar evaluando siguientes elementos si aplica
                            continue
                        except Exception:
                            continue
                    if dl_clicked_pre:
                        break
            except Exception:
                pass
        page.on("framenavigated", _nav_guard)
    except Exception:
        pass
    
    # Configuración avanzada de Playwright
    page.set_default_timeout(30000) # 30 segundos por defecto para todas las operaciones
    
    # Agregar simulación de movimientos de ratón aleatorios para parecer más humano
    page.evaluate('''
    () => {
      const events = ['click', 'touchstart', 'mouseover', 'mousemove'];
      events.forEach(event => {
        document.addEventListener(event, () => {}, true);
      });
    }
    ''')

    logging.info("→ Abriendo página...")
    try:
        # Simular comportamiento humano con intervalos aleatorios
        logging.info("Navegando directamente a la página de login del banco...")
        try:
            page.goto("https://particulares.bancosantander.es/login/", wait_until="domcontentloaded", timeout=60000)
        except Exception as e_goto1:
            logging.warning(f"Primer intento de goto(login) falló: {e_goto1}")
            # Reintento con menor condición de carga
            try:
                page.goto("https://particulares.bancosantander.es/login/", wait_until="commit", timeout=60000)
            except Exception as e_goto2:
                logging.error(f"Segundo intento de goto(login) falló: {e_goto2}")
                # Evidencias deshabilitadas
                pass
                raise
        
        login_start = time.time()
        logging.info("Página de login cargada")
        
        # Evidencias deshabilitadas
        
        # Intentar detectar redirecciones o cambios en la URL
        current_url = page.url
        logging.info(f"URL actual: {current_url}")
        
    except Exception as e:
        logging.error(f"Error al cargar la página: {e}")
        sys.exit(1)

    # Aceptar/ocultar cookies de forma robusta para no bloquear clics posteriores
    try:
        logging.info("Intentando aceptar/ocultar banner de cookies...")
        # 1) Esperar a que el banner de cookies esté presente
        try:
            page.wait_for_selector("#onetrust-banner-sdk, .onetrust-banner, .cookie-banner, .cookies-banner", timeout=10000)
            logging.info("Banner de cookies detectado")
        except Exception:
            logging.info("No se detectó banner de cookies o ya fue aceptado previamente")
        
        # 2) Intento directo por botones comunes de aceptación
        try:
            # Botones estándar de OneTrust y variantes
            accept_selectors = [
                "#onetrust-accept-btn-handler",
                "button#onetrust-accept-btn-handler",
                ".onetrust-banner-sdk-button.onetrust-accept-btn-handler",
                "#acceptAllCookies",
                "button[data-ot-accept-all]",
                "button.js-accept-all-cookies",
                "button.cookie-accept-all",
                "button.accept-cookies-button",
                "#cookies-accept-button",
                "button#cookies-accept-button"
            ]
            clicked = False
            for selector in accept_selectors:
                try:
                    if page.is_visible(selector):
                        # Esperar a que el botón sea clickable
                        page.wait_for_selector(f"{selector}:not([disabled])", timeout=5000)
                        page.click(selector, timeout=5000)
                        logging.info(f"Clic realizado en botón de aceptar cookies: {selector}")
                        clicked = True
                        break
                except Exception:
                    continue
            
            # Si no se encontró ningún botón, intentar con el botón genérico de OneTrust
            if not clicked:
                try:
                    if page.is_visible("#onetrust-accept-btn-handler"):
                        page.click("#onetrust-accept-btn-handler", timeout=5000)
                        logging.info("Clic realizado en botón de aceptar cookies (método alternativo)")
                        clicked = True
                except Exception:
                    pass
        except Exception:
            pass
        
        # 3) Llamada directa a APIs de gestión de cookies si existen
        try:
            page.evaluate("""
                () => {
                  try { if (window.OneTrust && OneTrust.AcceptAll) { OneTrust.AcceptAll(); } } catch(e){}
                  try { if (window.Cookiebot && Cookiebot.dialog && Cookiebot.dialog.submitConsent) { Cookiebot.dialog.submitConsent(); } } catch(e){}
                  try { if (window.Optanon && Optanon.ToggleInfoDisplay) { Optanon.ToggleInfoDisplay(); } } catch(e){}
                }
            """)
        except Exception:
            pass
        
        # 4) Ocultar forzosamente overlays y banners de cookies
        try:
            page.evaluate("""
                () => {
                  try {
                    const selectors = [
                        '#onetrust-consent-sdk',
                        '.onetrust-pc-dark-filter',
                        '.onetrust-banner',
                        '.cookie-banner',
                        '.cookies-banner',
                        '.js-cookie-banner',
                        '.ot-sdk-container',
                        '.optanon-alert-box-wrapper'
                    ];
                    selectors.forEach(sel => {
                        const el = document.querySelector(sel);
                        if (el) {
                            el.style.display='none';
                            el.style.pointerEvents='none';
                            el.style.visibility='hidden';
                        }
                    });
                  } catch(e){}
                }
            """)
        except Exception:
            pass
        
        # 5) Esperar a que el banner se cierre completamente
        try:
            page.wait_for_selector("#onetrust-banner-sdk, .onetrust-banner, .cookie-banner, .cookies-banner", state="detached", timeout=10000)
            logging.info("Banner de cookies cerrado correctamente")
        except Exception:
            logging.warning("No se pudo verificar el cierre completo del banner de cookies, continuando...")
        
        # 6) Verificación final
        page.wait_for_timeout(1000)
        try:
            overlay_present = page.evaluate("""
                () => {
                    const selectors = [
                        '#onetrust-consent-sdk .onetrust-pc-dark-filter',
                        '.onetrust-banner',
                        '.cookie-banner:not([style*="display: none"]):not([style*="visibility: hidden"])',
                        '.cookies-banner:not([style*="display: none"]):not([style*="visibility: hidden"])'
                    ];
                    return selectors.some(sel => document.querySelector(sel));
                }
            """)
        except Exception:
            overlay_present = False
        if overlay_present:
            logging.warning("Algunos elementos de cookies podrían seguir visibles; se ha forzado su ocultación.")
    except Exception as e:
        logging.warning(f"No se pudo gestionar el banner de cookies: {e}")
    
    # Esperar a estado de red estable tras gestionar cookies
    try:
        page.wait_for_load_state("networkidle")
    except Exception:
        pass
    # Reducir espera tras cookies para acelerar llegada al detalle
    page.wait_for_timeout(500)
    
    # Evidencias deshabilitadas tras aceptar cookies
    
    if time.time() - login_start > MAX_LOGIN_TIME:
        logging.error("Demasiado tiempo intentando hacer login. Reiniciando script...")
        python_exe = sys.executable
        os.execl(python_exe, python_exe, /home/sami/santander/venv/bin/python, *sys.argv)
    
    # Enumerar todos los campos de entrada visibles para analizar
    logging.info("Analizando todos los campos de entrada visibles en la página")
    try:
        # Obtener lista de todos los inputs en la página
        all_inputs = page.evaluate('''
            () => {
                const inputs = Array.from(document.querySelectorAll('input'));
                return inputs.map(input => {
                    return {
                        id: input.id || 'sin-id',
                        type: input.type || 'sin-tipo',
                        name: input.name || 'sin-nombre',
                        placeholder: input.placeholder || 'sin-placeholder',
                        class: input.className || 'sin-clase',
                        visible: input.offsetParent !== null
                    };
                });
            }
        ''')
        
        logging.info(f"Se encontraron {len(all_inputs)} campos de entrada")
        for i, inp in enumerate(all_inputs):
            logging.info(f"Input {i}: id={inp['id']}, type={inp['type']}, name={inp['name']}, visible={inp['visible']}")
    except Exception as e:
        logging.error(f"Error al analizar campos de entrada: {e}")
    
    # Rellenar NIF - probar diferentes selectores posibles
    # Ampliamos la lista de selectores posibles basados en la experiencia y posibles cambios en la web
    input_selectors = [
        "#inputDocuNumber",
        "input[name='documento']", 
        ".documento-input",
        "input[placeholder*='DNI']", 
        "input[placeholder*='documento']",
        "input[placeholder*='N.I.F']",
        "input[placeholder*='Documento']",
        "input[type='text']:visible",
        "#docNumber",
        "input.login-document",
        "input[name*='doc']",
        "input[name*='nif']",
        "input[name*='dni']",
        "input[id*='doc']",
        "input[id*='dni']",
        "input[id*='nif']",
        "input.form-control:visible"
    ]
                      
    selector_encontrado = False
    
    for selector in input_selectors:
        try:
            logging.info(f"Intentando encontrar el selector: {selector}")
            # Usamos una estrategia diferente: primero verificamos si existe y es visible
            exists = page.evaluate(f'''
                () => {{
                    const element = document.querySelector("{selector}");
                    if (element) {{
                        const style = window.getComputedStyle(element);
                        const isVisible = style.display !== 'none' && style.visibility !== 'hidden' && style.opacity !== '0';
                        return isVisible;
                    }}
                    return false;
                }}
            ''')
            
            if exists:
                logging.info(f"Selector encontrado y visible: {selector}")
                # (Sin PNG) Solo log informativo del selector encontrado
                logging.info("Selector visible, se continúa sin capturar PNG")
                
                # Simular comportamiento humano con tiempo de escritura
                page.click(selector)
                time.sleep(random.uniform(0.5, 1.2))
                
                # Escribir el NIF letra por letra con pequeños retrasos aleatorios
                for char in nif:
                    page.keyboard.type(char)
                    time.sleep(random.uniform(0.05, 0.2))
                
                selector_encontrado = True
                break
        except Exception as e:
            logging.warning(f"Error al interactuar con selector {selector}: {e}")
    
    if not selector_encontrado:
        logging.error("No se encontró ningún campo para introducir el NIF")
        # Evidencias deshabilitadas
        sys.exit(1)
        
    if time.time() - login_start > MAX_LOGIN_TIME:
        print("Demasiado tiempo intentando hacer login. Reiniciando script...")
        python_exe = sys.executable
        os.execl(python_exe, python_exe, /home/sami/santander/venv/bin/python, *sys.argv)

    # Esperar campos clave y rellenar solo los visibles
    page.wait_for_selector("#inputPW0", timeout=10000)
    for i in range(8):
        sel = f"#inputPW{i}"
        el = page.query_selector(sel)
        if el and el.is_visible() and el.is_enabled():
            page.click(sel)
            page.type(sel, clave[i])
        if time.time() - login_start > MAX_LOGIN_TIME:
            print("Demasiado tiempo intentando hacer login. Reiniciando script...")
            /home/sami/santander/venv/bin/python = sys.executable
            os.execl(python_exe, python_exe, /home/sami/santander/venv/bin/python, *sys.argv)

    # Pulsar botón Entrar
    page.locator("button:has-text('Entrar')").click()
    if time.time() - login_start > MAX_LOGIN_TIME:
        print("Demasiado tiempo intentando hacer login. Reiniciando script...")
        python_exe = sys.executable
        os.execl(python_exe, python_exe, /home/sami/santander/venv/bin/python, *sys.argv)

    # Cerrar ventana emergente si aparece
    try:
        page.locator("#mcp-cross-close").click(timeout=5000)
    except:
        pass

    # Permanecer en la página actual de cuenta; no navegar a otra URL
    page.wait_for_load_state("networkidle")
    page.wait_for_timeout(2000)

    # Intentar cerrar modal de filtros si aparece
    try:
        cerrar_btn = page.locator(".c-sanicon__close")
        if cerrar_btn.count() > 0 and cerrar_btn.first.is_visible():
            cerrar_btn.first.click()
            print("Modal de filtros cerrado")
    except Exception as e:
        print("No se pudo cerrar el modal de filtros (puede que no esté visible):", e)

    # Seleccionar la cuenta antes de aplicar filtros de fecha
    try:
        print("Esperando a que se muestren las tarjetas de cuentas ...")
        page.wait_for_selector(".san-product-cards__interactive-layer", timeout=15000)
        cuentas = page.locator(".san-product-cards__interactive-layer")
        print(f"Encontradas {cuentas.count()} cuentas en pantalla")
        if cuentas.count() > 0:
            cuenta_btn = cuentas.first
            if cuenta_btn.is_visible():
                print("Clickando en la primera cuenta disponible ...")
                # Esperar la navegación que provoca el click en la tarjeta de cuenta
                with page.expect_navigation(wait_until="networkidle"):
                    cuenta_btn.click()
                page.wait_for_timeout(3000)
                # Hacer clic en el breadcrumb para asegurarnos de estar en la vista correcta
                try:
                    breadcrumb = page.locator("div.flame-breadcrumbs").first
                    if breadcrumb.is_visible():
                        print("Clickando breadcrumb de navegación ...")
                        breadcrumb.click()
                        page.wait_for_timeout(2000)
                    else:
                        print("Breadcrumb de navegación no visible")
                except Exception as e:
                    print(f"Error al hacer clic en breadcrumb: {e}")
            else:
                print("La cuenta encontrada no está visible para hacer clic")
        else:
            print("No se encontraron cuentas para seleccionar")
    except Exception as e:
        print(f"No se pudo seleccionar la cuenta: {e}")

    # Entrar explícitamente en el detalle de la cuenta 'SAN. ONE EMPRESA' si no estamos ya en detalle
    try:
        # Comprobar si ya estamos en detalle por URL o hash
        is_detail = False
        try:
            cur_url = page.url
        except Exception:
            cur_url = ''
        try:
            cur_hash = page.evaluate("() => location.hash || ''")
        except Exception:
            cur_hash = ''
        is_detail = ('/oneweb/nhb/cuentas/detalle' in (cur_url or '')) or ('/cuentas/detalle' in (cur_hash or ''))
        if not is_detail:
            print("Buscando y accediendo a la cuenta 'SAN. ONE EMPRESA'...")
            page.wait_for_load_state("domcontentloaded")
            page.wait_for_timeout(800)
            import re as _re
            pattern = _re.compile(r"SAN\.?\s*ONE\s*EMPRESA", _re.I)
            # Intento 1: búsqueda directa en main/frame con scroll incremental
            def _try_find_and_click(ctx):
                try:
                    el = ctx.get_by_text(pattern).first
                    if el and el.count() > 0:
                        # Intentar clicar ancestro clicable cercano
                        anchor = el
                        try:
                            anchor = el.locator('xpath=ancestor-or-self::button|ancestor-or-self::a|ancestor::div[@role="button"][1]').first
                        except Exception:
                            anchor = el
                        try:
                            anchor.scroll_into_view_if_needed()
                        except Exception:
                            pass
                        try:
                            anchor.click()
                        except Exception:
                            try:
                                anchor.click(force=True)
                            except Exception:
                                el.click(force=True)
                        return True
                except Exception:
                    pass
                return False

            def _scan_with_scroll(ctx, max_steps=8, step=600):
                try:
                    # ir arriba
                    ctx.evaluate("() => window.scrollTo(0, 0)")
                except Exception:
                    pass
                for _ in range(max_steps):
                    if _try_find_and_click(ctx):
                        return True
                    try:
                        ctx.evaluate("(s) => window.scrollBy(0, s)", step)
                    except Exception:
                        pass
                    page.wait_for_timeout(250)
                return _try_find_and_click(ctx)

            clicked = False
            # Reintento: primero en main, luego en frames
            if _scan_with_scroll(page):
                clicked = True
            else:
                for fr in page.frames:
                    if fr == page.main_frame:
                        continue
                    if _scan_with_scroll(fr):
                        clicked = True
                        break
            # Reintento final: navegar a lista de cuentas y repetir
            if not clicked:
                try:
                    print("Reintentando: navegando a lista de cuentas...")
                    page.goto("https://particulares.bancosantander.es/oneweb/nhb/cuentas", wait_until="load")
                except Exception:
                    page.goto("https://particulares.bancosantander.es/oneweb/nhb/cuentas")
                page.wait_for_load_state("networkidle")
                page.wait_for_timeout(1000)
                if _scan_with_scroll(page):
                    clicked = True
            if clicked:
                # Esperar a que cargue el detalle
                try:
                    page.wait_for_function("() => (location.href||'').includes('/oneweb/nhb/cuentas/detalle') || (location.hash||'').includes('/cuentas/detalle')", timeout=30000)
                except Exception:
                    pass
                page.wait_for_load_state("networkidle")
                page.wait_for_timeout(1500)
                # --- NUEVO: Pulsar botón "Descargar" (abre modal de opciones) ---
                try:
                    print("Buscando botón 'Descargar' (wrap-chip) …")
                    modal_abierto = False
                    try:
                        # Intento 1: por rol y nombre accesible
                        btn_desc = page.get_by_role("button", name=re.compile("Descargar", re.I)).first
                        if btn_desc and btn_desc.count() > 0 and btn_desc.is_visible():
                            btn_desc.click()
                            # Esperar apertura de modal
                            page.wait_for_timeout(500)
                            try:
                                dlg = page.locator(":is([role=dialog], san-modal, .modal, .cdk-overlay-container, .mat-dialog-container)")
                                if dlg and dlg.count() > 0:
                                    print("Modal de opciones de descarga abierto (por rol)")
                                    modal_abierto = True
                            except Exception:
                                pass
                    except Exception:
                        pass
                    if not modal_abierto:
                        # Intento 2: selector específico del botón indicado
                        try:
                            btn_chip = page.locator("button.wrap-chip[aria-label*='Descargar' i]").first
                            if btn_chip and btn_chip.count() > 0 and btn_chip.is_visible():
                                btn_chip.click()
                                page.wait_for_timeout(500)
                                try:
                                    dlg = page.locator(":is([role=dialog], san-modal, .modal, .cdk-overlay-container, .mat-dialog-container)")
                                    if dlg and dlg.count() > 0:
                                        print("Modal de opciones de descarga abierto (wrap-chip)")
                                        modal_abierto = True
                                except Exception:
                                    pass
                        except Exception:
                            pass
                    if modal_abierto:
                        print("Botón 'Descargar' pulsado correctamente. Buscando opción Excel...")
                        page.wait_for_timeout(1000)  # Esperar que cargue el modal completamente
                        descargado = False
                        
                        # Lista de selectores para encontrar la opción Excel - Nuevo selector específico primero
                        excel_selectors = [
                            "button.san-selector.checkbox:has-text('Descargar Excel')",
                            "button[role=checkbox]:has-text('Descargar Excel')",
                            "button.san-selector:has(p:has-text('Descargar Excel'))",
                            "button:has(p:text('Descargar Excel'))",
                            "[class*='san-selector' i][role='checkbox']:has-text('Descargar Excel')",
                            "button:has-text('Excel')",
                            "button:has-text('XLS')",
                            "[role='button']:has-text('Excel')",
                            "a:has-text('Excel')",
                            ".modal button:has-text('Excel')"
                        ]
                        
                        # Buscar y hacer clic en opción de Excel
                        for selector in excel_selectors:
                            try:
                                excel_option = page.locator(selector).first
                                if excel_option and excel_option.count() > 0 and excel_option.is_visible():
                                    print(f"Opción Excel encontrada con selector: {selector}")
                                    
                                    # Primero hacer clic en el checkbox sin esperar descarga
                                    print("Marcando checkbox de Excel...")
                                    try:
                                        # Intentar varios métodos para marcar el checkbox
                                        try:
                                            excel_option.check()
                                        except Exception:
                                            try:
                                                # Usar la API de Playwright en lugar de evaluate con querySelectorAll
                                                if 'has-text' in selector:
                                                    # Usar un selector más simple
                                                    simple_sel = "button:has-text('Excel')"
                                                    page.locator(simple_sel).check()
                                                else:
                                                    # Intentar clic por JavaScript en una ubicación relativa
                                                    box = excel_option.bounding_box()
                                                    if box:
                                                        page.mouse.click(box['x'] + box['width']/2, box['y'] + box['height']/2)
                                            except Exception:
                                                # Intentar clic por JavaScript en una ubicación relativa
                                                box = excel_option.bounding_box()
                                                if box:
                                                    page.mouse.click(box['x'] + box['width']/2, box['y'] + box['height']/2)
                                        except Exception:
                                            # Intentar clic por JavaScript en una ubicación relativa
                                            box = excel_option.bounding_box()
                                            if box:
                                                page.mouse.click(box['x'] + box['width']/2, box['y'] + box['height']/2)
                                        # Asegurarse que el checkbox está marcado usando JavaScript sin pasar el selector
                                        try:
                                            # Usar una forma alternativa más compatible con cualquier tipo de checkbox
                                            page.evaluate("""
                                            () => {
                                              // Buscar todos los checkboxes visibles en el modal
                                              const checkboxes = Array.from(document.querySelectorAll('input[type="checkbox"]'));
                                              const labels = Array.from(document.querySelectorAll('label'));
                                              
                                              // Buscar por texto cercano a "Excel"
                                              let found = false;
                                              
                                              // 1. Buscar en checkboxes directamente
                                              for (const cb of checkboxes) {
                                                const label = cb.parentElement?.textContent || '';
                                                if (label.toLowerCase().includes('excel')) {
                                                  cb.checked = true;
                                                  cb.dispatchEvent(new Event('change', {bubbles: true}));
                                                  found = true;
                                                  break;
                                                }
                                              }
                                              
                                              // 2. Buscar en labels si no se encontró
                                              if (!found) {
                                                for (const label of labels) {
                                                  if (label.textContent.toLowerCase().includes('excel')) {
                                                    const input = label.querySelector('input[type="checkbox"]');
                                                    if (input) {
                                                      input.checked = true;
                                                      input.dispatchEvent(new Event('change', {bubbles: true}));
                                                      found = true;
                                                      break;
                                                    }
                                                  }
                                                }
                                              }
                                              
                                              return found;
                                            }
                                            """)
                                        except Exception as e:
                                            print(f"Error al intentar marcar checkbox con JavaScript puro: {e}")
                                            pass
                                        
                                        # Esperar a que reaccione la interfaz
                                        page.wait_for_timeout(2000)
                                        
                                        # Ahora buscar el botón de aceptar/confirmar/descargar final
                                        print("Buscando botón de confirmación para iniciar descarga...")
                                        confirm_selectors = [
                                            # Selector exacto del botón proporcionado
                                            "button.san-ending-button.primary.large[aria-label='Descargar documentos seleccionados']",
                                            "button.san-ending-button.primary",
                                            "button[aria-label='Descargar documentos seleccionados']",
                                            "button[aria-description*='descargar']",
                                            # Otros selectores genéricos como fallback
                                            "button:has-text('Aceptar')", 
                                            "button:has-text('Confirmar')",
                                            "button:has-text('Descargar')",
                                            "button.primary"
                                        ]
                                        
                                        confirm_clicked = False
                                        for confirm_sel in confirm_selectors:
                                            try:
                                                btn = page.locator(confirm_sel).first
                                                if btn and btn.count() > 0 and btn.is_visible():
                                                    print(f"Encontrado botón de confirmación: {confirm_sel}")
                                                    
                                                    # Preparar para capturar el evento de descarga
                                                    with page.expect_event("download", timeout=30000) as dl_info:
                                                        # Hacer clic en el botón de confirmación
                                                        print(f"Haciendo clic en botón: {confirm_sel}")
                                                        btn.click()
                                                        page.wait_for_timeout(2000)
                                                    
                                                    # Procesar descarga
                                                    download = dl_info.value
                                                    try:
                                                        import os as _os
                                                        # Guardar directamente en /tmp
                                                        destino = "/tmp/descarga.xlsx"
                                                        download.save_as(destino)
                                                        print(f"✓ Descarga Excel guardada como: {destino}")
                                                        descargado = True
                                                        confirm_clicked = True
                                                        break
                                                    except Exception as e:
                                                        print(f"Error al guardar descarga: {e}")
                                                        try:
                                                            path = download.path()
                                                            print(f"✓ Descarga completada en ruta temporal: {path}")
                                                            descargado = True
                                                            confirm_clicked = True
                                                            break
                                                        except:
                                                            print("No se pudo obtener la ruta de la descarga")
                                            except Exception as e:
                                                print(f"Error con selector {confirm_sel}: {e}")
                                        
                                        # Si no se encontró/hizo clic en ningún botón, intentar con el checkbox directo
                                        if not confirm_clicked:
                                            try:
                                                with page.expect_event("download", timeout=10000) as dl_info:
                                                    excel_option.click()
                                                    page.wait_for_timeout(2000)
                                                    
                                                # Procesar descarga
                                                download = dl_info.value
                                                import os as _os
                                                _os.makedirs('/var/www/html/descargas', exist_ok=True)
                                                sugerido = download.suggested_filename or download.suggested_filename()
                                                destino = f"/var/www/html/descargas/{sugerido}"
                                                download.save_as(destino)
                                                print(f"✓ Descarga Excel (alternativa) guardada en: {destino}")
                                                descargado = True
                                            except Exception as e:
                                                print(f"No se pudo descargar directamente: {e}")
                                    except Exception as e:
                                        print(f"Error al procesar modal/confirmación: {e}") 
                            except Exception as e:
                                print(f"Error al intentar opción Excel {selector}: {e}")
                        
                        if not descargado:
                            print("No se pudo seleccionar opción Excel en el modal. Intentando alternativas...")
                            
                            # Intentar con elementos genéricos que puedan contener "Excel"
                            try:
                                # Buscar cualquier elemento visible con texto "Excel" 
                                excel_items = page.locator(":is(button,a,li,span,div):visible:has-text('Excel')").all()
                                if excel_items and len(excel_items) > 0:
                                    for item in excel_items[:5]:  # Limitar a los primeros 5
                                        try:
                                            with page.expect_event("download", timeout=10000) as dl_info:
                                                print("Intentando clic en elemento con texto Excel...")
                                                item.click(timeout=3000)
                                            
                                            download = dl_info.value
                                            sugerido = download.suggested_filename or download.suggested_filename()
                                            destino = f"/var/www/html/descargas/{sugerido}"
                                            download.save_as(destino)
                                            print(f"Descarga Excel (alternativo) guardada en: {destino}")
                                            descargado = True
                                            break
                                        except Exception:
                                            continue
                            except Exception as e:
                                print(f"Error en búsqueda alternativa de Excel: {e}")
                        
                        # Si no se logró con el modal, proceder con otros métodos
                        if not descargado:
                            print("No se logró descargar Excel desde el modal. Intentando fallback...")
                    else:
                        print("No se detectó el modal tras pulsar 'Descargar'. Se intentará descarga directa (fallback)…")
                    
                    # --- Fallback anterior: intentar botón/icono que dispare directamente una descarga ---
                    print("Buscando botón/icono de descarga para descarga directa…")
                    destinos = [page]
                    try:
                        for fr in page.frames:
                            if fr != page.main_frame:
                                destinos.append(fr)
                    except Exception:
                        pass
                    sel_lista = [
                        "button:has-text('Descargar')",
                        "button:has-text('Download')",
                        "[role='button']:has-text('Descargar')",
                        "a[download]",
                        "button[aria-label*='Descargar' i]",
                        "button[title*='Descargar' i]",
                        "[data-test*='download' i]",
                        "[data-testid*='download' i]",
                        "[class*='download' i]",
                        "svg[*,*]>>xpath=ancestor-or-self::button|ancestor-or-self::a[contains(@class,'download') or contains(@aria-label,'Descargar') or contains(@title,'Descargar')]"
                    ]
                    descargado = False
                    for ctx in destinos:
                        for sel in sel_lista:
                            try:
                                loc = ctx.locator(sel).first
                                if loc and loc.count() > 0 and loc.first.is_visible():
                                    print(f"Intentando clic de descarga con selector: {sel}")
                                    with page.expect_event("download", timeout=30000) as dl_info:
                                        try:
                                            loc.first.click(timeout=5000)
                                        except Exception:
                                            loc.first.click(force=True, timeout=5000)
                                    download = dl_info.value
                                    try:
                                        import os as _os
                                        _os.makedirs('/var/www/html/descargas', exist_ok=True)
                                        sugerido = download.suggested_filename or download.suggested_filename()
                                        destino = f"/var/www/html/descargas/{sugerido}"
                                        download.save_as(destino)
                                        print(f"Descarga guardada en: {destino}")
                                    except Exception:
                                        path = download.path()
                                        print(f"Descarga completada. Ruta temporal: {path}")
                                    descargado = True
                                    break
                            except Exception:
                                continue
                        if descargado:
                            break
                    if not descargado:
                        print("No se encontró botón de descarga visible. Reintentando con búsqueda heurística…")
                        try:
                            # Heurística: buscar elementos con texto/aria/ttl que contengan 'Descarg'
                            btns = page.locator(":is(button,a,[role=button],[aria-label],[title])").all()
                        except Exception:
                            btns = []
                        for b in btns[:20]:
                            try:
                                meta = b.evaluate("el => ({t:(el.innerText||'').trim(), a:el.getAttribute('aria-label')||'', tt:el.getAttribute('title')||''})")
                            except Exception:
                                meta = {"t":"","a":"","tt":""}
                            hay = any(s for s in [meta.get('t',''), meta.get('a',''), meta.get('tt','')] if 'descarg' in s.lower())
                            if hay:
                                print("Intentando clic heurístico de descarga…")
                                try:
                                    with page.expect_event("download", timeout=30000) as dl_info:
                                        b.click()
                                    d2 = dl_info.value
                                    sugerido = d2.suggested_filename or d2.suggested_filename()
                                    dest = f"/var/www/html/descargas/{sugerido}"
                                    d2.save_as(dest)
                                    print(f"Descarga guardada en: {dest}")
                                    descargado = True
                                    break
                                except Exception:
                                    continue
                    if not descargado:
                        print("No fue posible realizar la descarga automáticamente.")
                except Exception as e:
                    print(f"Error al intentar descargar: {e}")
            else:
                print("No se encontró el elemento con texto 'SAN. ONE EMPRESA'.")
    except Exception as e:
        print(f"Error al intentar entrar al detalle de la cuenta SAN. ONE EMPRESA: {e}")

    # Estado de navegación actual (solo informativo)
    try:
        try:
            cur_url = page.url
        except Exception:
            cur_url = "(sin url)"
        try:
            cur_hash = page.evaluate("() => location.hash || ''")
        except Exception:
            cur_hash = "(sin hash)"
        print(f"Estado de navegación actual: url={cur_url} hash={cur_hash}")
    except Exception as e:
        print(f"Error en verificación de estado de detalle: {e}")

    # Preparar ámbitos de búsqueda (página principal y frames), sin depender de textos específicos
    scope = page
    scopes = [page]
    try:
        for fr in page.frames:
            if fr != page.main_frame:
                scopes.append(fr)
    except Exception:
        pass

    # Evidencias deshabilitadas en vista de detalle

    # Detección del icono de descarga (solo resaltar, sin clics ni descargas)
    mk = None
    for scope in scopes:
        try:
            print("Iniciando búsqueda del botón de descarga...")
            res = scope.evaluate(
                r"""
                () => {
                  console.log('Iniciando búsqueda del botón de descarga...');
                  const pageRoot = document.querySelector('main') || document.body;
                  console.log('pageRoot encontrado:', pageRoot);
                  const actionContainers = Array.from(pageRoot.querySelectorAll('.buttons, .actions, .toolbar, .filters, .panel, .wrapper, [role="toolbar"], .san-actions, .san-toolbar, .table-actions, .actions-bar, .filters__actions, .san-section__actions, .san-card__actions'))
                    .filter(el => !el.closest('.header__buttons, header, nav, san-sidebar-menu, #personalAreaMenu, #mailboxMenu, #helpMenu, .menu, [role="menu"]'));
                  console.log('actionContainers encontrados:', actionContainers.length);
                  const dlSelParts = [
                    'a[aria-label*="Descargar" i]','button[aria-label*="Descargar" i]','[aria-label*="Exportar" i]','[title*="Descargar" i]','[title*="Exportar" i]',
                    'button:has(san-icon[iconcontent*="download" i])','a:has(san-icon[iconcontent*="download" i])','san-icon[iconcontent*="download" i]',
                    'button:has(san-icon[iconcontent*="arrow-down" i])','a:has(san-icon[iconcontent*="arrow-down" i])','san-icon[iconcontent*="arrow-down" i]',
                    'em.c-sanicon__download',
                    '[data-dtname*="download" i]','[data-dtname*="export" i]','[data-testid*="export" i]','[data-icon*="download" i]',
                    '[class*="download" i]','[class*="export" i]','[class*="descarg" i]',
                    'button[aria-label*="descargar" i]','a[aria-label*="descargar" i]',
                    '.flame-icon-download',
                    'san-chip#downloadPill'
                  ];
                  const collectSel = dlSelParts.join(',');
                  console.log('Selectores a buscar:', collectSel);
                  const all = Array.from(pageRoot.querySelectorAll(collectSel)).filter(el => {
                    if (!el) return false;
                    const anc = el.closest('header, nav, san-sidebar-menu, #personalAreaMenu, #mailboxMenu, #helpMenu, .menu, [role="menu"]');
                    if (anc) return false;
                    const hrefEl = (el.tagName && el.tagName.toLowerCase()==='a') ? el : el.closest('a');
                    const href = hrefEl && hrefEl.getAttribute && (hrefEl.getAttribute('href')||'');
                    if (href && /asuntos-clientes|buzon|ayuda|gestiones/i.test(href)) return false;
                    const st = getComputedStyle(el);
                    if (st.display==='none' || st.visibility==='hidden' || parseFloat(st.opacity||'1')===0) return false;
                    return true;
                  });
                  console.log('Elementos encontrados:', all.length);
                  const cont = actionContainers.find(c => c.querySelector(collectSel)) || pageRoot;
                  let cand = all.find(el => cont.contains(el)) || all[0] || null;
                  console.log('Candidato encontrado:', cand);
                  if (!cand) {
                    console.log('No se encontró candidato para el botón de descarga');
                    return { mark: null };
                  }
                  if (cand && cand.tagName && cand.tagName.toLowerCase()==='san-icon') {
                    const wrap = cand.closest('a,button,[role=button]');
                    if (wrap) cand = wrap;
                  }
                  const mark = 'cascade_'+Date.now();
                  try { cand.setAttribute('data-cascade-click', mark); } catch(e) {}
                  try { cand.scrollIntoView({block:'center', inline:'center'}); } catch(e) {}
                  console.log('Botón de descarga marcado con ID:', mark);
                  return { mark };
                }
                """
            )
            print(f"Resultado de la búsqueda: {res}")
            if isinstance(res, dict) and res.get('mark'):
                mk = res.get('mark')
                print(f"Botón de descarga encontrado y marcado con ID: {mk}")
                # Aplicar realce visual persistente (sin hacer clic)
                try:
                    scope.evaluate(
                        """
                        sel => {
                          const t = document.querySelector(sel);
                          if(!t) return false;
                          try { t.scrollIntoView({block:'center', inline:'center'}); } catch(e){}
                          // Estilos persistentes hasta el fin del script
                          try { t.style.setProperty('transition','box-shadow .2s ease, outline .2s ease','important'); } catch(e){}
                          try { t.style.setProperty('outline','3px solid #1e90ff','important'); } catch(e){}
                          try { t.style.setProperty('box-shadow','0 0 0 4px rgba(30,144,255,.25), 0 0 12px 2px rgba(30,144,255,.6)','important'); } catch(e){}
                          return true;
                        }
                        """,
                        f"[data-cascade-click='{mk}']"
                    )
                    
                    # Resaltar el icono marcado y hacer clic
                    try:
                        print(f"Buscando elemento con selector: [data-cascade-click='{mk}']")
                        marked_element = scope.locator(f"[data-cascade-click='{mk}']").first
                        print(f"Elemento encontrado: {marked_element is not None}")
                        if marked_element and marked_element.count() > 0:
                            print("Elemento marcado encontrado, verificando visibilidad...")
                            # Verificar si el elemento es visible
                            try:
                                is_visible = marked_element.is_visible()
                                print(f"Elemento es visible: {is_visible}")
                            except Exception as visibility_err:
                                print(f"Error al verificar visibilidad: {visibility_err}")
                            
                            # Asegurar que el elemento esté visible
                            try:
                                marked_element.scroll_into_view_if_needed()
                                print("Icono de descarga: resaltado exitosamente.")
                            except Exception as scroll_err:
                                print(f"Error al hacer scroll al elemento: {scroll_err}")
                            
                            # Pequeña pausa para observar el resaltado
                            page.wait_for_timeout(2000)
                            
                            # Hacer clic en el botón de descarga
                            try:
                                print("Intentando hacer clic en el botón de descarga...")
                                marked_element.click()
                                print("Icono de descarga: clic realizado exitosamente.")
                                
                                # Esperar a que aparezca la modal
                                print("Esperando a que aparezca la modal...")
                                page.wait_for_timeout(3000)
                                print("Tiempo de espera completado. Buscando modal...")
                                
                                # Verificar si hay una modal visible
                                try:
                                    modal_elements = page.locator(".modal, [role='dialog'], [aria-modal='true']").all()
                                    print(f"Elementos de modal encontrados: {len(modal_elements)}")
                                    for i, modal in enumerate(modal_elements[:3]):
                                        try:
                                            if modal.is_visible():
                                                print(f"Modal {i+1} es visible")
                                            else:
                                                print(f"Modal {i+1} no es visible")
                                        except Exception as modal_check_err:
                                            print(f"Error al verificar visibilidad del modal {i+1}: {modal_check_err}")
                                except Exception as modal_check_err:
                                    print(f"Error al buscar modales: {modal_check_err}")
                                
                                # Buscar y hacer clic en el botón "Descargar Excel" en la modal
                                try:
                                    # Esperar un poco más para que la modal se cargue completamente
                                    page.wait_for_timeout(2000)
                                    
                                    # Imprimir información de depuración sobre los elementos visibles
                                    print("Buscando botón 'Descargar Excel' en la modal...")
                                    
                                    # Selector para el botón "Descargar Excel" en la modal
                                    download_excel_button = page.locator("text=Descargar Excel").first
                                    if download_excel_button and download_excel_button.count() > 0:
                                        download_excel_button.click()
                                        print("Botón 'Descargar Excel': clic realizado exitosamente.")
                                    else:
                                        # Intentar con otros selectores comunes
                                        selectors = [
                                            "button:has-text('Descargar Excel')",
                                            "[aria-label*='Descargar Excel' i]",
                                            ".modal button:has-text('Descargar Excel')",
                                            "button:has-text('Descargar excel')",  # Variación en minúsculas
                                            "button:has-text('descargar excel')",  # Todo en minúsculas
                                            "[data-testid*='download' i]",
                                            "[class*='download' i]",
                                            "[class*='export' i]",
                                            "[class*='descarg' i]",
                                            "button:has(span:text('Descargar Excel'))",
                                            "button:has(div:text('Descargar Excel'))"
                                        ]
                                        found = False
                                        for selector in selectors:
                                            try:
                                                button = page.locator(selector).first
                                                if button and button.count() > 0 and button.is_visible():
                                                    button.click()
                                                    print(f"Botón 'Descargar Excel' (selector {selector}): clic realizado exitosamente.")
                                                    found = True
                                                    break
                                                elif button and button.count() > 0:
                                                    print(f"Botón encontrado con selector '{selector}' pero no es visible.")
                                            except Exception as selector_err:
                                                print(f"Error al buscar con selector '{selector}': {selector_err}")
                                                continue
                                        
                                        if not found:
                                            # Si no se encuentra el botón, mostrar un mensaje de advertencia
                                            print("Advertencia: No se encontró el botón 'Descargar Excel' en la modal.")
                                            # Guardar una evidencia adicional para depuración
                                            try:
                                                page.screenshot(path="/home/sami/logs/modal_debug.png", full_page=True)
                                                print("Evidencia adicional guardada en: /home/sami/logs/modal_debug.png")
                                                
                                                # Imprimir información adicional sobre los elementos en la página
                                                try:
                                                    buttons = page.locator("button").all()
                                                    print(f"Total de botones encontrados en la página: {len(buttons)}")
                                                    for i, btn in enumerate(buttons[:10]):  # Solo los primeros 10 para no sobrecargar
                                                        try:
                                                            text = btn.text_content()
                                                            if text and len(text.strip()) > 0:
                                                                print(f"  Botón {i+1}: '{text.strip()}'")
                                                        except:
                                                            pass
                                                except Exception as enum_err:
                                                    print(f"Error al enumerar botones: {enum_err}")
                                            except Exception as screenshot_err:
                                                print(f"No se pudo guardar la evidencia adicional: {screenshot_err}")
                                except Exception as modal_err:
                                    print(f"Error al hacer clic en 'Descargar Excel': {modal_err}")
                            except Exception as click_err:
                                print(f"Icono de descarga: error al hacer clic: {click_err}")
                        else:
                            print("No se encontró el elemento marcado para el botón de descarga.")
                    except Exception as highlight_err:
                        print(f"Icono de descarga: error al resaltar: {highlight_err}")
                    
                    break
                except Exception as e:
                    print(f"Error al procesar el icono de descarga: {e}")
            else:
                print("No se encontró botón de descarga para marcar.")
        except Exception as _e:
            print(f"Error en descarga previa (scope): {_e}")

    # Guardar evidencia HTML con el botón resaltado
    try:
        import os
        from datetime import datetime
        # Crear directorio para evidencias si no existe
        evidencias_dir = "/home/sami/logs"
        os.makedirs(evidencias_dir, exist_ok=True)
        
        # Generar nombre de archivo con timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        html_path = f"{evidencias_dir}/evidencia_boton_descarga_{timestamp}.html"
        
        # Obtener el HTML completo de la página
        html_content = page.content()
        
        # Guardar el HTML
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print(f"Evidencia HTML guardada en: {html_path}")
    except Exception as e:
        print(f"Error al guardar evidencia HTML: {e}")
    
    # Mantener el navegador abierto para observar el resaltado
    try:
        print("Botón de descarga resaltado. Manteniendo navegador abierto para observación...")
        page.wait_for_timeout(30000)  # Mantener abierto por 30 segundos para observación
    except Exception:
        pass
    
    # Cerrar navegador y finalizar
    try:
        context.close()
    except Exception:
        try:
            browser.close()
        except Exception:
            pass
    print("Navegador cerrado. Script finalizado.")
    
    # Llamar al script scrapeo.py al finalizar
    try:
        import subprocess
        print("Ejecutando scrapeo.py...")
                subprocess.call(["/home/sami/santander/venv/bin/python", "/home/sami/santander/scrapeo.py"])
    except Exception as e:
        print(f"Error al ejecutar scrapeo.py: {e}")

